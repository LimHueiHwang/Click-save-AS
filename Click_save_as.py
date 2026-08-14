import datetime
import os
import sys
import traceback

import cv2
import numpy as np
import pyautogui
import win32com.client as win32
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from utils.constant import (
    MAX_RETRIES,
    error_file,
    target_wb_name,
)


# ============================================================
# CONFIGURATION
# ============================================================

MATCH_THRESHOLD = 0.8
SCALE_RANGE = np.linspace(0.8, 1.2, 10)
DEBUG_MODE = False


# ============================================================
# EXCEL ERROR WORKBOOK
# ============================================================

def close_error_file(file_path, retries=5):
    """
    Close the error workbook if it is currently open in Excel.

    The workbook is closed without saving because the error log
    is written separately by the automation.
    """
    if not file_path:
        return

    error_file_abs = os.path.abspath(file_path).lower()

    try:
        excel = win32.GetActiveObject("Excel.Application")
    except Exception:
        return

    for _ in range(retries):
        workbook_found = False

        try:
            workbooks = list(excel.Workbooks)
        except Exception:
            return

        for workbook in workbooks:
            try:
                workbook_path = os.path.abspath(workbook.FullName).lower()

                if workbook_path == error_file_abs:
                    workbook_found = True
                    workbook.Close(SaveChanges=False)
                    break

            except Exception:
                continue

        if not workbook_found:
            return


# ============================================================
# IMAGE RECOGNITION
# ============================================================

def capture_screen():
    """Capture the current screen and return it as a BGR OpenCV image."""
    screenshot = pyautogui.screenshot()

    return cv2.cvtColor(
        np.array(screenshot),
        cv2.COLOR_RGB2BGR,
    )


def load_template(image_path):
    """Load a reference image and return it as a grayscale image."""
    template = cv2.imread(image_path)

    if template is None:
        print(f"[Warning] Unable to load image: {image_path}")
        return None

    return cv2.cvtColor(
        template,
        cv2.COLOR_BGR2GRAY,
    )


def find_best_match(screen_gray, template_gray):
    """
    Find the best multi-scale template match on the screen.

    Returns:
        tuple: (confidence, location, scale)
    """
    template_height, template_width = template_gray.shape[:2]

    best_value = 0
    best_location = None
    best_scale = 1

    for scale in SCALE_RANGE:
        resized_template = cv2.resize(
            template_gray,
            None,
            fx=scale,
            fy=scale,
            interpolation=cv2.INTER_LINEAR,
        )

        resized_height, resized_width = resized_template.shape[:2]

        if (
            resized_height > screen_gray.shape[0]
            or resized_width > screen_gray.shape[1]
        ):
            continue

        result = cv2.matchTemplate(
            screen_gray,
            resized_template,
            cv2.TM_CCOEFF_NORMED,
        )

        _, max_value, _, max_location = cv2.minMaxLoc(result)

        if max_value > best_value:
            best_value = max_value
            best_location = max_location
            best_scale = scale

    return (
        best_value,
        best_location,
        best_scale,
        template_width,
        template_height,
    )


def calculate_match_center(
    location,
    template_width,
    template_height,
    scale,
):
    """Calculate the center coordinates of a matched template."""
    top_left = location

    bottom_right = (
        top_left[0] + int(template_width * scale),
        top_left[1] + int(template_height * scale),
    )

    center_x = (
        top_left[0]
        + (bottom_right[0] - top_left[0]) // 2
    )

    center_y = (
        top_left[1]
        + (bottom_right[1] - top_left[1]) // 2
    )

    return top_left, bottom_right, center_x, center_y


def click_position(center_x, center_y):
    """Move the mouse to the specified position and click."""
    pyautogui.moveTo(
        center_x,
        center_y,
        duration=0.3,
    )

    pyautogui.click()


def find_and_click_icon_multiscale(
    image_path,
    threshold=MATCH_THRESHOLD,
):
    """
    Locate an icon on the current screen using multi-scale
    OpenCV template matching and click its center.

    Returns:
        True  - icon detected and clicked
        False - icon not found or detection failed
    """
    if not image_path or not os.path.exists(image_path):
        return False

    try:
        screen_img = capture_screen()

        screen_gray = cv2.cvtColor(
            screen_img,
            cv2.COLOR_BGR2GRAY,
        )

        template_gray = load_template(image_path)

        if template_gray is None:
            return False

        (
            best_value,
            best_location,
            best_scale,
            template_width,
            template_height,
        ) = find_best_match(
            screen_gray,
            template_gray,
        )

        if best_value < threshold or best_location is None:
            print(
                f"🔍 Icon not found for "
                f"{os.path.basename(image_path)} | "
                f"Best confidence: {best_value:.2f}"
            )
            return False

        (
            top_left,
            bottom_right,
            center_x,
            center_y,
        ) = calculate_match_center(
            best_location,
            template_width,
            template_height,
            best_scale,
        )

        if DEBUG_MODE:
            debug_img = screen_img.copy()

            cv2.rectangle(
                debug_img,
                top_left,
                bottom_right,
                (0, 0, 255),
                2,
            )

            cv2.imshow("Detected Icon", debug_img)
            cv2.waitKey(0)
            cv2.destroyAllWindows()

        click_position(
            center_x,
            center_y,
        )

        print(
            f"✅ Found icon {image_path} at "
            f"({center_x},{center_y}) | "
            f"Confidence: {best_value:.2f}"
        )

        return True

    except Exception as error:
        print(
            f"[Error] find_and_click_icon_multiscale: {error}"
        )
        return False


# ============================================================
# IMAGE LOCATION
# ============================================================

def find_mode_images():
    """
    Locate available Dark Mode and Light Mode Save As reference images.

    The automation checks the local Desktop folder first and then
    the OneDrive Desktop folder. The first folder containing at
    least one reference image is used.
    """
    user_profile = os.environ.get("USERPROFILE")

    if not user_profile:
        print(
            "[Error] USERPROFILE environment variable not found."
        )
        return []

    folders = [
        os.path.join(
            user_profile,
            "Desktop",
            "PO Softcopy",
        ),
        os.path.join(
            user_profile,
            "OneDrive - Jabil",
            "Desktop",
            "PO Softcopy",
        ),
    ]

    image_names = [
        "Dark_Mode.png",
        "Light_Mode.png",
    ]

    for folder in folders:
        image_paths = []

        for image_name in image_names:
            image_path = os.path.join(
                folder,
                image_name,
            )

            if os.path.exists(image_path):
                image_paths.append(image_path)

        if image_paths:
            return image_paths

    print(
        "[Error] No Save As reference images were found."
    )

    return []


# ============================================================
# SAP SAVE AS PROCESS
# ============================================================

def process_po(po_number):
    """
    Detect and click the SAP Save As icon.

    Available Dark Mode and Light Mode reference images are
    tried on each retry.

    The PO number is used for error reporting. The surrounding
    VBA workflow remains responsible for the overall PO process.
    """
    image_paths = find_mode_images()

    if not image_paths:
        raise FileNotFoundError(
            "Unable to locate any Save As reference images."
        )

    for attempt in range(1, MAX_RETRIES + 1):
        print(
            f"Attempt {attempt}/{MAX_RETRIES} "
            "to locate SAP 'Save As' button..."
        )

        for image_path in image_paths:
            if find_and_click_icon_multiscale(image_path):
                return

    raise RuntimeError(
        "Failed to find SAP 'Save As' button "
        f"after {MAX_RETRIES} retries for PO {po_number}."
    )


# ============================================================
# ERROR LOGGING
# ============================================================

def log_error(po_number, error_text):
    """
    Write an automation error to the Excel error log.
    """
    if not error_file:
        return

    if os.path.exists(error_file):
        workbook = load_workbook(error_file)
    else:
        workbook = Workbook()

    if "Errors" in workbook.sheetnames:
        worksheet = workbook["Errors"]
    else:
        worksheet = workbook.active
        worksheet.title = "Errors"

        worksheet.append(
            ["PO Number", "Time", "Error"]
        )

    time_str = datetime.datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    worksheet.append(
        [
            po_number,
            time_str,
            error_text,
        ]
    )

    worksheet[f"C{worksheet.max_row}"].alignment = Alignment(
        wrap_text=True
    )

    workbook.save(error_file)


# ============================================================
# PO NUMBER
# ============================================================

def get_po_number():
    """
    Retrieve the PO number passed from the VBA macro.
    """
    if len(sys.argv) <= 1:
        raise ValueError("Unable to find PO number.")

    return sys.argv[1]

def normalize_po_number(value):
    """Convert an Excel or command-line PO value to an integer."""
    return int(float(value))


# ============================================================
# VBA WORKBOOK ERROR UPDATE
# ============================================================

def log_error_macro(po_number, error_text):
    """
    Write the error message to the corresponding PO row
    in the running Excel macro workbook.
    """
    try:
        excel = win32.GetActiveObject(
            "Excel.Application"
        )

        target_workbook = next(
            (
                workbook
                for workbook in excel.Workbooks
                if workbook.Name == target_wb_name
            ),
            None,
        )

        if not target_workbook:
            return

        worksheet = target_workbook.Sheets("Macro")
        current_po = normalize_po_number(po_number)

        row = 4

        while worksheet.Cells(row, 1).Value is not None:
            cell_po = normalize_po_number(
                worksheet.Cells(row, 1).Value
            )

            if cell_po == current_po:
                worksheet.Cells(row, 4).Value = error_text
                break

            row += 1

    except Exception as error:
        print(
            f"[Warning] Unable to update VBA workbook: {error}"
        )


# ============================================================
# MAIN
# ============================================================

def main():
    close_error_file(error_file)

    po_number = "UNKNOWN"

    try:
        po_number = get_po_number()
        process_po(po_number)

    except Exception as error:
        clean_error = (
            f"{type(error).__name__}: {error}"
        )

        print(clean_error)

        try:
            log_error(
                po_number,
                clean_error,
            )
        except Exception as log_error_exception:
            print(
                f"[Warning] Unable to write error log: "
                f"{log_error_exception}"
            )

        try:
            log_error_macro(
                po_number,
                clean_error,
            )
        except Exception as macro_error:
            print(
                f"[Warning] Unable to update VBA workbook: "
                f"{macro_error}"
            )


if __name__ == "__main__":
    main()
